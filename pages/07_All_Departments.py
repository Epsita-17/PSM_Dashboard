import io
import re
from html.parser import HTMLParser
from datetime import datetime

import pandas as pd
import requests
from openpyxl import load_workbook
import streamlit as st
import plotly.graph_objects as go

# ============================================================
# PAGE CONFIG
# ============================================================
st.set_page_config(
    page_title="PSM Dashboard - All Departments",
    page_icon="🏭",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ============================================================
# GOOGLE SHEET
# ============================================================
SPREADSHEET_ID = "1--X0TT5Ts92EKAxrhV-fQgqeTHBX3rDVc1Egg74MewM"

# IMPORTANT:
# These are the GIDs used for the All Departments dashboard.
# PT and the other modules are NOT read by row count from the
# whole workbook. Each module is loaded from its own tab/GID.
SHEETS = {
    "PT": "1997330551",
    "PHA": "1151637695",
    "PHA Recommendation": "1114420199",
    "MOC": "1493447251",
    "PSSR": "1914804736",
    "PS Incident": "354502422",  # corrected: Incident
    "Training": "1071736559",  # corrected: Training
    "SOC-SOL": "510439154",
    "Critical Equipment": None,
    "Alarm": None,
    "Barrier Audit": "1790395364",
}

# ============================================================
# STYLE
# ============================================================
st.markdown(
    """
    <style>

    .stApp {
        background:#f3f8fc;
    }

    #MainMenu, footer {
        visibility:hidden;
    }



   .block-container {
    padding:0rem 0.35rem 0rem 0.35rem !important;
    margin-bottom:0px !important;
    max-width:100%;
}

/* REMOVE BOTTOM SPACE */
[data-testid="stAppViewContainer"] {
    padding-bottom:0px !important;
}

[data-testid="stMainBlockContainer"] {
    padding-bottom:0px !important;
    margin-bottom:0px !important;
}

section.main {
    padding-bottom:0px !important;
    margin-bottom:0px !important;
}

   div[data-testid="stMetric"] {
    background:#ffffff;
    border:1px solid #cbddea;
    border-radius:7px;
    padding:6px 6px !important;
    min-height:72px;
    overflow:visible !important;
}

    div[data-testid="stMetricLabel"] {
    font-size:9px !important;
    font-weight:800 !important;
    color:#20384f !important;
    white-space:nowrap !important;
    overflow:visible !important;
    text-overflow:clip !important;
    line-height:1.1 !important;
}
div[data-testid="stMetricLabel"],
div[data-testid="stMetricLabel"] > div,
div[data-testid="stMetricLabel"] p {
    overflow:visible !important;
    text-overflow:clip !important;
    white-space:nowrap !important;
    max-width:none !important;
}

div[data-testid="stMetricLabel"] p {
    margin:0 !important;
    padding:0 !important;
    font-size:8px !important;
    line-height:1.1 !important;
}
    div[data-testid="stMetricValue"] {
        color:#123f77 !important;
        font-size:24px !important;
        font-weight:900 !important;
    }

    .module-card {
        background:#ffffff;
        border:1px solid #d3e0ea;
        border-radius:7px;
        padding:7px;
        margin-bottom:8px;
        box-shadow:0 1px 4px rgba(20,65,95,.06);
    }

    .module-title {
        color:#073f78;
        font-size:12px;
        font-weight:900;
        margin-bottom:6px;
    }

    .section-bar {
        background:#07518b;
        color:#ffffff;
        border-radius:4px;
        padding:6px 8px;
        font-size:10px;
        font-weight:900;
        margin:4px 0 6px 0;
    }

    .live-bar {
        background:#ffffff;
        border:1px solid #cbddea;
        border-radius:4px;
        padding:5px 8px;
        color:#4f6678;
        font-size:10px;
        margin-bottom:6px;
    }

    .footer {
        text-align:center;
        color:#627689;
        background:#edf4f8;
        border-top:1px solid #cbdce7;
        padding:7px;
        font-size:10px;
        font-weight:800;
        margin-top:8px;
    }

    .small-note {
        font-size:9px;
        color:#6c7f8f;
    }

    .stDataFrame {
        border:1px solid #d5e0e8;
    }

    /* ============================================================
       AUDIT REGISTER - COMPACT LOOK
       ============================================================ */
    .audit-register-wrap {
        width:100%;
        max-height:180px;
        overflow-y:auto;
        overflow-x:hidden;
        border:1px solid #d5e0e8;
        border-radius:4px;
        background:#ffffff;
    }

    .audit-register-table {
        width:100%;
        border-collapse:collapse;
        table-layout:fixed;
        font-size:10px;
        color:#173f70;
    }

    .audit-register-table th {
        position:sticky;
        top:0;
        z-index:2;
        background:#f3f7fa;
        color:#627689;
        font-weight:800;
        text-align:left;
        padding:7px 8px;
        border-bottom:1px solid #d5e0e8;
    }

    .audit-register-table td {
        padding:7px 8px;
        border-bottom:1px solid #e1e8ee;
        white-space:nowrap;
        overflow:hidden;
        text-overflow:ellipsis;
    }

    .audit-register-table th:nth-child(1), .audit-register-table td:nth-child(1) { width:8%; }
    .audit-register-table th:nth-child(2), .audit-register-table td:nth-child(2) { width:24%; }
    .audit-register-table th:nth-child(3), .audit-register-table td:nth-child(3) { width:18%; }
    .audit-register-table th:nth-child(4), .audit-register-table td:nth-child(4) { width:17%; }
    .audit-register-table th:nth-child(5), .audit-register-table td:nth-child(5) { width:33%; }

    .audit-register-table a {
        color:#0067c5 !important;
        font-weight:700;
        text-decoration:none;
    }

    .audit-register-table a:hover {
        text-decoration:underline;
    }


    /* ========================================================
       REDUCE SPACE BETWEEN HEADER AND REFRESH BUTTON
       ======================================================== */

    div[data-testid="stButton"] {
        margin-top:-35px !important;
        margin-bottom:0px !important;
    }


/* ============================================================
   MATCH CLICKABLE MODULE HEADINGS WITH NORMAL MODULE HEADING
   ============================================================ */
[data-testid="stPageLink"] {
    margin-bottom:6px !important;
}

[data-testid="stPageLink"],
[data-testid="stPageLink"] a,
[data-testid="stPageLink"] a *,
[data-testid="stPageLink"] p,
[data-testid="stPageLink"] span,
[data-testid="stPageLink"] div {
    color:#073f78 !important;
    font-size:12px !important;
    font-weight:900 !important;
    text-decoration:none !important;
}

[data-testid="stPageLink"] a:hover,
[data-testid="stPageLink"] a:hover * {
    color:#073f78 !important;
    text-decoration:none !important;
}

</style>
    """,
    unsafe_allow_html=True,
)

# ALL DEPARTMENTS HEADER
# Same tested header design; only department title is changed
# ============================================================

header_html = """
<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">

<style>

* {
    box-sizing: border-box;
}

html, body {
    margin: 0;
    padding: 0;
    width: 100%;
    height: 100%;
    overflow: hidden;
    font-family: Arial, Helvetica, sans-serif;
}

body {
    background: #f4f9fc;
}

.header {
    position: relative;
    width: 100%;
    height: 145px;
    overflow: hidden;

    display: flex;
    align-items: center;
    justify-content: center;

    background:
        radial-gradient(
            ellipse at center,
            rgba(55,160,218,.24) 0%,
            rgba(223,242,252,.82) 45%,
            rgba(244,250,253,.98) 100%
        ),
        linear-gradient(
            180deg,
            #edf8fd 0%,
            #dceff8 100%
        );

    border-top: 2px solid #0b91d1;
    border-bottom: 3px solid #1487c2;

    box-shadow:
        0 4px 12px rgba(21,92,130,.18);
}

.header::before {
    content: "";
    position: absolute;
    inset: 0;

    background-image:
        radial-gradient(
            circle,
            rgba(0,122,190,.17) 1.2px,
            transparent 1.5px
        );

    background-size: 15px 15px;
    opacity: .65;
}

.header::after {
    content: "";
    position: absolute;
    inset: 0;

    background:
        linear-gradient(
            135deg,
            transparent 0 7%,
            rgba(0,133,210,.12) 7% 8%,
            transparent 8% 11%,
            rgba(0,133,210,.08) 11% 12%,
            transparent 12%
        ),
        linear-gradient(
            315deg,
            transparent 0 7%,
            rgba(0,133,210,.12) 7% 8%,
            transparent 8% 11%,
            rgba(0,133,210,.08) 11% 12%,
            transparent 12%
        );
}

.content {
    position: relative;
    z-index: 8;

    width: 100%;
    height: 100%;

    display: flex;
    flex-direction: column;

    align-items: center;
    justify-content: center;

    text-align: center;
}

.title {
    color: #153e68;
    font-size: 24px;
    font-weight: 950;
    letter-spacing: 5px;
    line-height: 1;
    margin-bottom: 5px;

    text-shadow:
        0 1px 1px rgba(255,255,255,.9);
}

.pillar {
    position: relative;

    width: 560px;
    height: 66px;

    display: flex;
    align-items: center;
    justify-content: center;

    background:
        linear-gradient(
            180deg,
            #176ca5 0%,
            #07518b 55%,
            #063e70 100%
        );

    border: 1px solid #0877ba;
    border-radius: 14px;

    color: #ffd21a;

    font-size: 42px;
    font-weight: 950;
    letter-spacing: 1px;

    box-shadow:
        0 7px 16px rgba(11,83,130,.25),
        inset 0 1px 0 rgba(255,255,255,.28),
        inset 0 -5px 12px rgba(0,35,75,.16);
}

.pillar::before,
.pillar::after {
    position: absolute;

    top: 50%;
    transform: translateY(-50%);

    color: #51c5ff;
    font-size: 21px;
    font-weight: 950;
    letter-spacing: -5px;

    text-shadow:
        0 1px 5px rgba(0,100,160,.5);
}

.pillar::before {
    content: "◀◀";
    left: 17px;
}

.pillar::after {
    content: "▶▶";
    right: 17px;
}

.subtitle {
    margin-top: 7px;

    height: 25px;
    min-width: 700px;

    display: flex;
    align-items: center;
    justify-content: center;

    padding: 4px 30px;

    background:
        linear-gradient(
            90deg,
            #075b8e,
            #1188c4,
            #075b8e
        );

    border: 1px solid #078fd2;
    border-radius: 7px;

    color: #ffffff;

    font-size: 10px;
    font-weight: 900;
    letter-spacing: 1.8px;

    box-shadow:
        0 4px 9px rgba(10,93,140,.20),
        inset 0 1px 0 rgba(255,255,255,.25);
}

.top-line {
    position: absolute;

    top: 0;
    left: 24%;

    width: 52%;
    height: 3px;

    background:
        linear-gradient(
            90deg,
            transparent,
            #00a9ff 18%,
            #ffffff 50%,
            #00a9ff 82%,
            transparent
        );

    box-shadow:
        0 0 8px rgba(0,169,255,.55);
}

.corner-light {
    position: absolute;
    z-index: 10;

    width: 110px;
    height: 3px;

    background:
        linear-gradient(
            90deg,
            transparent,
            #00baff,
            transparent
        );

    box-shadow:
        0 0 8px rgba(0,186,255,.55);
}

.corner-left {
    left: 7%;
    top: 7px;
}

.corner-right {
    right: 7%;
    top: 7px;
}

</style>
</head>

<body>

<div class="header">

    <div class="top-line"></div>

    <div class="corner-light corner-left"></div>
    <div class="corner-light corner-right"></div>

    <div class="content">

        <div class="title">
            PSM DASHBOARD
        </div>

        <div class="pillar">
            ALL DEPARTMENTS
        </div>

        <div class="subtitle">
            PROCESS SAFETY MANAGEMENT DIGITAL VISION WALL
        </div>

    </div>

</div>

</body>
</html>
"""

st.components.v1.html(
    header_html,
    height=170,
    scrolling=False
)


# ============================================================
# HELPERS
# ============================================================
def norm(value):
    """Normalize text for safe column/department matching."""
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())


def find_col(df, candidates):
    """Find a column using exact normalized match, then partial match."""
    if df is None or df.empty:
        return None

    normalized = {norm(c): c for c in df.columns}

    for candidate in candidates:
        key = norm(candidate)
        if key in normalized:
            return normalized[key]

    for column in df.columns:
        ckey = norm(column)
        for candidate in candidates:
            nkey = norm(candidate)
            if nkey in ckey or ckey in nkey:
                return column

    return None


def clean_dataframe(df):
    if df is None or df.empty:
        return pd.DataFrame()

    out = df.copy()
    out.columns = [
        str(c).strip().replace("\n", " ").replace("\r", " ")
        for c in out.columns
    ]

    out = out.dropna(how="all").copy()

    for c in out.columns:
        if out[c].dtype == "object":
            out[c] = out[c].astype(str).str.strip()

    return out


# ============================================================
# GLOBAL DEPARTMENT FILTER
# ============================================================

def filter_selected_department(df, selected_department):
    df = clean_dataframe(df)

    if df.empty:
        return df

    # ALL DEPARTMENTS = no filtering
    if selected_department == "All Departments":
        return df.copy()

    department_col = find_col(
        df,
        [
            "Department",
            "Departments",
            "Dept",
            "Department Name",
            "Department_Name",
            "Dept Name",
            "Dept_Name",
        ],
    )

    # If department column does not exist,
    # return empty for an individual department.
    if department_col is None:
        return pd.DataFrame(columns=df.columns)

    department_values = (
        df[department_col]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    # Normalize both sides
    selected_norm = norm(selected_department)

    data_norm = department_values.map(norm)

    # Normal exact match
    mask = data_norm == selected_norm

    # Additional handling for common department names
    if selected_norm == "blastfurnace":
        mask = department_values.str.contains(
            r"blast\s*[-_/ ]*\s*furnace",
            case=False,
            regex=True,
            na=False,
        )

    elif selected_norm == "cokeoven":
        mask = department_values.str.contains(
            r"coke\s*[-_/ ]*oven",
            case=False,
            regex=True,
            na=False,
        )


    elif selected_norm == "sms1":
        mask = department_values.str.contains(
            r"sms\s*[-_/ ]*1",
            case=False,
            regex=True,
            na=False,
        )

    elif selected_norm == "sms2":
        mask = department_values.str.contains(
            r"sms\s*[-_/ ]*2",
            case=False,
            regex=True,
            na=False,
        )

    elif selected_norm == "centralutility":
        mask = department_values.str.contains(
            r"central\s*[-_/ ]*utility",
            case=False,
            regex=True,
            na=False,
        )

    elif selected_norm == "tubeMill".lower():
        mask = department_values.str.contains(
            r"tube\s*[-_/ ]*mill",
            case=False,
            regex=True,
            na=False,
        )

    elif selected_norm == "pelletbeneficiation":
        mask = department_values.str.contains(
            r"pellet.*beneficiation|beneficiation.*pellet",
            case=False,
            regex=True,
            na=False,
        )

    return df.loc[mask].copy()


def load_csv_from_url(url):
    response = requests.get(
        url,
        timeout=30,
        headers={"User-Agent": "Mozilla/5.0"},
    )
    response.raise_for_status()

    # Google export is normally UTF-8 CSV.
    # Use StringIO so pandas gets a text stream.
    return pd.read_csv(io.StringIO(response.content.decode("utf-8-sig")))


@st.cache_data(ttl=300, show_spinner=False)
def load_google_sheet(gid):
    """
    Load ONE exact Google Sheet tab by GID.

    This avoids the previous problem where the code used sheet
    names and could end up reading the wrong tab.
    """
    if not gid:
        return pd.DataFrame()

    export_url = (
        f"https://docs.google.com/spreadsheets/d/"
        f"{SPREADSHEET_ID}/export?format=csv&gid={gid}"
    )

    try:
        df = load_csv_from_url(export_url)
        return clean_dataframe(df)
    except Exception:
        # Fallback to Google visualization endpoint.
        gviz_url = (
            f"https://docs.google.com/spreadsheets/d/"
            f"{SPREADSHEET_ID}/gviz/tq?"
            f"tqx=out:csv&gid={gid}"
        )

        try:
            df = load_csv_from_url(gviz_url)
            return clean_dataframe(df)
        except Exception:
            return pd.DataFrame()


class _AuditLinkParser(HTMLParser):
    """Extract hyperlinks from the Compliance Report column of gviz HTML."""
    def __init__(self):
        super().__init__()
        self.rows = []
        self._row = None
        self._cell = None
        self._cell_href = None
        self._in_table = False

    def handle_starttag(self, tag, attrs):
        attrs = dict(attrs)
        if tag == "table":
            self._in_table = True
        elif self._in_table and tag == "tr":
            self._row = []
        elif self._in_table and tag in ("td", "th") and self._row is not None:
            self._cell = True
            self._cell_href = None
        elif self._in_table and tag == "a" and self._cell:
            self._cell_href = attrs.get("href")

    def handle_endtag(self, tag):
        if tag in ("td", "th") and self._cell:
            self._row.append(self._cell_href or "")
            self._cell = None
            self._cell_href = None
        elif tag == "tr" and self._row is not None:
            self.rows.append(self._row)
            self._row = None
        elif tag == "table":
            self._in_table = False


def load_audit_with_links(gid):
    """Load Audit tab and dynamically extract the current Compliance Report hyperlinks.

    The URL is NEVER hard-coded. Each refresh downloads the current Google Sheet
    XLSX and reads the hyperlink attached to the Compliance Report cell.
    This means a new monthly report/link is picked up automatically.
    """
    df = load_google_sheet(gid)
    if df.empty or not gid:
        return df

    url_values = [""] * len(df)

    xlsx_url = (
        f"https://docs.google.com/spreadsheets/d/"
        f"{SPREADSHEET_ID}/export?format=xlsx&gid={gid}"
    )

    try:
        response = requests.get(
            xlsx_url,
            timeout=30,
            headers={"User-Agent": "Mozilla/5.0"},
        )
        response.raise_for_status()

        workbook = load_workbook(
            filename=io.BytesIO(response.content),
            data_only=False,
            read_only=False,
        )

        # Google normally exports the requested GID as the active/first sheet.
        ws = workbook.active

        # Find the header row and Compliance Report column dynamically.
        header_row = None
        report_col = None

        for row in ws.iter_rows():
            for cell in row:
                value = str(cell.value or "").strip().lower()
                if value in {
                    "compliance report",
                    "compliance report link",
                    "audit report",
                    "report",
                }:
                    header_row = cell.row
                    report_col = cell.column
                    break
            if report_col is not None:
                break

        if report_col is not None:
            # Extract both direct cell hyperlinks and HYPERLINK formulas.
            hyperlink_rows = []

            for row_no in range((header_row or 1) + 1, ws.max_row + 1):
                cell = ws.cell(row=row_no, column=report_col)
                url = ""

                if cell.hyperlink:
                    try:
                        url = str(cell.hyperlink.target or "").strip()
                    except Exception:
                        url = ""

                # Also support =HYPERLINK("URL","Displayed Text")
                if not url and isinstance(cell.value, str):
                    formula = cell.value.strip()
                    match = re.search(
                        r'HYPERLINK\s*\(\s*["\'](https?://[^"\']+)["\']',
                        formula,
                        flags=re.IGNORECASE,
                    )
                    if match:
                        url = match.group(1).strip()

                hyperlink_rows.append(url)

            # Map by row order to the CSV records.
            for i in range(min(len(url_values), len(hyperlink_rows))):
                url_values[i] = hyperlink_rows[i]

        workbook.close()

    except Exception:
        # Do not break the dashboard if XLSX hyperlink extraction fails.
        pass

    # Last fallback: direct URL already present in CSV cell.
    fallback_col = find_col(
        df,
        [
            "Compliance Report",
            "Compliance Report Link",
            "Audit Report",
            "Report",
        ],
    )

    if fallback_col:
        for i in range(len(df)):
            if url_values[i]:
                continue
            value = str(df.iloc[i][fallback_col]).strip()
            if re.match(r"^https?://", value, re.I):
                url_values[i] = value

    df = df.copy()
    df["__COMPLIANCE_REPORT_URL"] = url_values
    return df


def load_module(name):
    """Load the complete module data for all departments."""
    return load_google_sheet(SHEETS.get(name))


def status_counts(df, status_candidates=None):
    result = {
        "total": 0,
        "completed": 0,
        "ongoing": 0,
        "pending": 0,
        "overdue": 0,
        "open": 0,
        "closed": 0,
    }

    if df is None or df.empty:
        return result

    result["total"] = len(df)

    if status_candidates is None:
        status_candidates = [
            "Status",
            "Current Status",
            "Action Status",
            "Completion Status",
            "Investigation Status",
            "Recommendation Status",
        ]

    status_col = find_col(df, status_candidates)

    if status_col is None:
        return result

    s = (
        df[status_col]
        .fillna("")
        .astype(str)
        .str.strip()
        .str.lower()
    )

    result["completed"] = int(
        s.str.contains(r"\bcompleted?\b|\bcomplete\b", regex=True).sum()
    )

    result["ongoing"] = int(
        s.str.contains(r"ongoing|in progress|in-progress", regex=True).sum()
    )

    result["pending"] = int(
        s.str.contains(r"pending", regex=True).sum()
    )

    result["overdue"] = int(
        s.str.contains(r"overdue", regex=True).sum()
    )

    result["open"] = int(
        s.str.fullmatch(r"open", case=False, na=False).sum()
    )

    result["closed"] = int(
        s.str.fullmatch(r"closed", case=False, na=False).sum()
    )

    return result


def make_register(df, id_names, description_names, status_names):
    if df is None or df.empty:
        return pd.DataFrame()

    id_col = find_col(df, id_names)
    desc_col = find_col(df, description_names)
    status_col = find_col(df, status_names)

    result = pd.DataFrame(index=df.index)

    # -----------------------------
    # ID / NUMBER
    # -----------------------------
    if id_col:
        result["No."] = (
            df[id_col]
            .fillna("")
            .astype(str)
            .str.strip()
        )
    else:
        result["No."] = [
            f"{i + 1:03d}" for i in range(len(df))
        ]

    # -----------------------------
    # DESCRIPTION
    # -----------------------------
    if desc_col:
        result["Description"] = (
            df[desc_col]
            .fillna("-")
            .astype(str)
            .str.strip()
        )
    else:
        result["Description"] = "-"

    # -----------------------------
    # STATUS
    # -----------------------------
    if status_col:
        result["Status"] = (
            df[status_col]
            .fillna("-")
            .astype(str)
            .str.strip()
        )
    else:
        result["Status"] = "-"

    return result.reset_index(drop=True)


# ============================================================
# STATUS COLOUR
# ============================================================

def status_style(value):
    value = str(value).strip().lower()

    if value in ["completed", "complete", "closed"]:
        return "color: #008000; font-weight: 800;"

    elif value in ["ongoing", "in progress", "in-progress", "pending"]:
        return "color: #e67e00; font-weight: 800;"

    elif value in ["overdue", "open"]:
        return "color: #d71920; font-weight: 800;"

    else:
        return "color: #333333;"


# ============================================================
# REGISTER DISPLAY
# ============================================================

def show_register(title, df, id_names, description_names, status_names):
    st.markdown(
        f'<div class="section-bar">{title}</div>',
        unsafe_allow_html=True,
    )

    register_df = make_register(
        df,
        id_names,
        description_names,
        status_names,
    )

    if register_df.empty:

        st.info("No records found.")

    else:

        # Apply colour ONLY to Status column
        styled_df = (
            register_df.style
            .map(
                status_style,
                subset=["Status"]
            )
        )

        st.dataframe(
            styled_df,
            use_container_width=True,
            hide_index=True,
            height=260,
        )



def make_moc_register(df):
    """Build the compact MOC register: MOC No., Description, Type, Status, Remarks."""
    if df is None or df.empty:
        return pd.DataFrame()

    moc_no_col = find_col(
        df,
        [
            "MOC No",
            "MOC No.",
            "MOC Number",
            "MOC ID",
            "Request No",
            "Request ID",
            "Sr No",
            "Sr. No",
            "Serial No",
            "ID",
        ],
    )

    desc_col = find_col(
        df,
        [
            "Description of Change",
            "Description of the Change",
            "MOC Description",
            "Change Description",
            "Description",
        ],
    )

    type_col = find_col(
        df,
        [
            "Change Type (Permanent/Temporary/Emergency)",
            "Change Type (Permanent / Temporary / Emergency)",
            "Change Type",
            "MOC Change Type",
            "Type of Change",
            "Type",
        ],
    )

    status_col = find_col(
        df,
        [
            "Status (Open/Close)",
            "Status (Open / Close)",
            "Status",
            "Current Status",
            "MOC Status",
        ],
    )

    remarks_col = find_col(
        df,
        [
            "Remarks",
            "Remark",
            "Comments",
            "Comment",
        ],
    )

    result = pd.DataFrame(index=df.index)

    if moc_no_col:
        result["MOC No."] = (
            df[moc_no_col].fillna("").astype(str).str.strip()
        )
    else:
        result["MOC No."] = [
            f"{i + 1:03d}" for i in range(len(df))
        ]

    result["Description"] = (
        df[desc_col].fillna("-").astype(str).str.strip()
        if desc_col else "-"
    )

    result["Type"] = (
        df[type_col].fillna("-").astype(str).str.strip()
        if type_col else "-"
    )

    result["Status"] = (
        df[status_col].fillna("-").astype(str).str.strip()
        if status_col else "-"
    )

    result["Remarks"] = (
        df[remarks_col].fillna("-").astype(str).str.strip()
        if remarks_col else "-"
    )

    return result.reset_index(drop=True)


def show_moc_register(df):
    st.markdown(
        '<div class="section-bar">MOC REGISTER</div>',
        unsafe_allow_html=True,
    )

    register_df = make_moc_register(df)

    if register_df.empty:
        st.info("No MOC records found.")
        return

    styled_df = (
        register_df.style
        .map(
            status_style,
            subset=["Status"]
        )
    )

    st.dataframe(
        styled_df,
        use_container_width=True,
        hide_index=True,
        height=120,
        column_config={
            "MOC No.": st.column_config.TextColumn("MOC No.", width="small"),
            "Description": st.column_config.TextColumn("Description", width="small"),
            "Type": st.column_config.TextColumn("Type", width="small"),
            "Status": st.column_config.TextColumn("Status", width="small"),
            "Remarks": st.column_config.TextColumn("Remarks", width="small"),
        },
    )



# ============================================================
# MODULE PAGE LINKS
# ============================================================

MODULE_PAGE_LINKS = {
    "PROCESS TECHNOLOGY (PT)": "pages/09_PT.py",
    "PROCESS HAZARD ANALYSIS (PHA)": "pages/10_PHA.py",
    "PHA RECOMMENDATION": "pages/10_PHA.py",
    "MOC": "pages/11_MOC.py",
    "PRE-STARTUP SAFETY REVIEW (PSSR)": "pages/12_PSSR.py",
    "PROCESS SAFETY INCIDENT": "pages/14_PSI.py",
    "TRAINING": "pages/13_Training.py",
}


def show_module_title(number, icon, title):
    page = MODULE_PAGE_LINKS.get(title)

    if page:
        st.page_link(
            page,
            label=f"🔴 {number} {icon} {title}",
        )
    else:
        st.markdown(
            f'<div class="module-title">🔴 {number} {icon} {title}</div>',
            unsafe_allow_html=True,
        )


def show_metric_row(items):
    cols = st.columns(len(items), gap="small")
    for c, (label, value) in zip(cols, items):
        with c:
            st.metric(label, value)


def get_date_column(df):
    return find_col(
        df,
        [
            "Date",
            "Audit Date",
            "Last Audit Date",
            "Deviation Date",
            "Record Date",
            "Incident Date",
        ],
    )


# ============================================================
# LOAD DATA
# ============================================================
if st.button("↻ Refresh Data", key="refresh_data"):
    st.cache_data.clear()
    st.rerun()

loaded = {}

for module_name in SHEETS:
    loaded[module_name] = load_module(module_name)

pt = loaded["PT"]
pha = loaded["PHA"]
rec = loaded["PHA Recommendation"]
moc = loaded["MOC"]
pssr = loaded["PSSR"]
training = loaded["Training"]
soc = loaded["SOC-SOL"]
incident = loaded["PS Incident"]
audit = load_audit_with_links(SHEETS["Barrier Audit"])

# ============================================================
# COMMON DEPARTMENT SELECTOR
# ============================================================

ALL_DEPARTMENTS = [
    "All Departments",
    "Blast Furnace",
    "Coke Oven",
    "SMS-1",
    "SMS-2",
    "DRI",
    "Central Utility",
    "CRM",
    "WRM",
    "CPP",
    "Sinter",
    "Tube Mill",
    "CSP",
    "Pellet & Beneficiation",
    "LCP",
]

selected_department = st.selectbox(
    "Department Status",
    ALL_DEPARTMENTS,
    index=0,
    key="global_department_selector",
)
# ============================================================
# APPLY GLOBAL DEPARTMENT FILTER
# ============================================================

pt = filter_selected_department(
    pt,
    selected_department
)

pha = filter_selected_department(
    pha,
    selected_department
)

rec = filter_selected_department(
    rec,
    selected_department
)

moc = filter_selected_department(
    moc,
    selected_department
)

pssr = filter_selected_department(
    pssr,
    selected_department
)

training = filter_selected_department(
    training,
    selected_department
)

soc = filter_selected_department(
    soc,
    selected_department
)

incident = filter_selected_department(
    incident,
    selected_department
)
# ============================================================
# LIVE DATA BAR
# ============================================================
module_count = sum(
    1 for df in loaded.values()
    if df is not None and not df.empty
)

st.markdown(
    f"""
    <div class="live-bar">
        <b>LIVE DATA: Google Sheet → All Departments</b>
        &nbsp; | &nbsp;
        Refresh: 5 minutes
        &nbsp; | &nbsp;
        Modules with data: <b>{module_count}</b>
        &nbsp; | &nbsp;
        Last load: <b>{datetime.now().strftime("%d-%b-%Y %H:%M:%S")}</b>
    </div>
    """,
    unsafe_allow_html=True,
)

# ============================================================
# ROW 1 — PT / PHA / RECOMMENDATION / MOC
# ============================================================
a, b, c, d = st.columns(4, gap="small")

with a:
    with st.container(
            border=True,
            height=460
    ):
        # PT HEADER
        show_module_title(
            1,
            "",
            "PROCESS TECHNOLOGY (PT)"
        )

        # PT KPI
        x = status_counts(pt)

        show_metric_row([
            ("TOTAL PT", x["total"]),
            ("COMPLETED", x["completed"]),
            ("ONGOING", x["ongoing"]),
        ])

        # PT REGISTER
        show_register(
            "PT REGISTER",
            pt,
            ["PT No", "PT No.", "PT ID", "ID"],
            ["PT Description", "Description", "PT Name"],
            ["Status", "Current Status"],
        )

# ============================================================
# 2 — PROCESS HAZARD ANALYSIS (PHA)
# ============================================================


with b:
    with st.container(border=True):
        show_module_title(
            2,
            "△",
            "PROCESS HAZARD ANALYSIS (PHA)"
        )

        x = status_counts(pha)

        show_metric_row([
            ("TOTAL PHA", x["total"]),
            ("COMPLETED", x["completed"]),
            ("ONGOING", x["ongoing"]),
        ])

        show_register(
            "PHA REGISTER",
            pha,
            [
                "PHA No",
                "PHA No.",
                "PHA ID",
                "ID"
            ],
            [
                "PHA Description",
                "Description",
                "PHA Name"
            ],
            [
                "Status",
                "Current Status"
            ],
        )

# ============================================================
# 3 — PHA RECOMMENDATION
# ============================================================

with c:
    with st.container(border=True):
        show_module_title(
            3,
            "♧",
            "PHA RECOMMENDATION"
        )

        x = status_counts(
            rec,
            [
                "Status (Open/Close)",
                "Status Open Close",
                "Open/Close Status",
                "Recommendation Status",
            ],
        )

        show_metric_row([
            (
                "TOTAL RECOMMENDATIONS",
                x["total"]
            ),
            (
                "OPEN",
                x["open"]
            ),
            (
                "CLOSED",
                x["closed"]
            ),
        ])

        show_register(
            "RECOMMENDATION REGISTER",
            rec,
            [
                "PHA No",
                "PHA No.",
                "Recommendation No",
                "Recommendation ID",
                "ID"
            ],
            [
                "Recommendation Description",
                "Recommendation",
                "Description"
            ],
            [
                "Status (Open/Close)",
                "Status Open Close",
                "Open/Close Status",
                "Recommendation Status",
                "Status"
            ],
        )

# ============================================================
# 4 — MANAGEMENT OF CHANGE (MOC)
# ============================================================

with d:
    with st.container(
            border=True,
            height=460
    ):

        show_module_title(
            3,
            "♙",
            "MOC"
        )

        x = status_counts(moc)

        show_metric_row([
            ("TOTAL MOC", x["total"]),
            ("OPEN", x["open"]),
            ("CLOSED", x["closed"]),
        ])

        # ----------------------------------------------------
        # MOC COLUMN MAPPING
        # ----------------------------------------------------

        moc_change_type_col = find_col(
            moc,
            [
                "Change Type (Permanent/Temporary/Emergency)",
                "Change Type",
                "Type",
            ],
        )

        moc_category_col = find_col(
            moc,
            [
                "Category of changes (Technology/Personnel/Facility)",
                "Category of changes",
                "Category",
            ],
        )

        moc_chart = moc.copy()

        # ----------------------------------------------------
        # ALL DEPARTMENTS
        # ----------------------------------------------------
        # No department filter is applied.
        # The complete MOC sheet is used.

        # ----------------------------------------------------
        # TWO DONUT CHARTS
        # ----------------------------------------------------

        p1, p2 = st.columns(
            2,
            gap="small"
        )

        # ====================================================
        # TYPE-WISE
        # ====================================================

        with p1:

            st.markdown(
                """
                <div style="
                    text-align:center;
                    color:#173f70;
                    font-size:10px;
                    font-weight:800;
                    margin-bottom:4px;
                ">
                    TYPE-WISE DISTRIBUTION
                </div>
                """,
                unsafe_allow_html=True,
            )

            if (
                    moc_change_type_col
                    and not moc_chart.empty
            ):

                type_data = (
                    moc_chart[moc_change_type_col]
                    .fillna("")
                    .astype(str)
                    .str.strip()
                )

                type_data = type_data[
                    type_data != ""
                    ]

                type_counts = (
                    type_data.value_counts()
                )

                if not type_counts.empty:

                    type_labels = [
                        f"{label} {int(value)} "
                        f"({value / type_counts.sum() * 100:.0f}%)"
                        for label, value
                        in zip(
                            type_counts.index,
                            type_counts.values
                        )
                    ]

                    fig_type = go.Figure(
                        data=[
                            go.Pie(
                                labels=type_labels,
                                values=type_counts.values,
                                hole=0.58,
                                textinfo="none",
                                domain=dict(
                                    x=[0.00, 0.55],
                                    y=[0.08, 0.92],
                                ),
                                hovertemplate=(
                                    "%{label}"
                                    "<extra></extra>"
                                ),
                            )
                        ]
                    )

                    fig_type.update_layout(
                        height=115,
                        margin=dict(
                            l=0,
                            r=0,
                            t=0,
                            b=0
                        ),
                        showlegend=True,
                        legend=dict(
                            orientation="v",
                            x=0.72,
                            y=0.5,
                            xanchor="left",
                            yanchor="middle",
                            font=dict(size=8),
                        ),
                        font=dict(size=8),
                    )

                    st.plotly_chart(
                        fig_type,
                        use_container_width=True,
                        config={
                            "displayModeBar": False
                        },
                        key="bf_moc_type_donut",
                    )

                else:

                    st.info(
                        "No MOC Type data found."
                    )

            else:

                st.info(
                    "MOC Change Type column not found."
                )

        # ====================================================
        # CATEGORY-WISE
        # ====================================================

        with p2:

            st.markdown(
                """
                <div style="
                    text-align:center;
                    color:#173f70;
                    font-size:10px;
                    font-weight:800;
                    margin-bottom:4px;
                ">
                    CATEGORY-WISE DISTRIBUTION
                </div>
                """,
                unsafe_allow_html=True,
            )

            if (
                    moc_category_col
                    and not moc_chart.empty
            ):

                category_data = (
                    moc_chart[moc_category_col]
                    .fillna("")
                    .astype(str)
                    .str.strip()
                )

                category_data = category_data[
                    category_data != ""
                    ]

                category_counts = (
                    category_data.value_counts()
                )

                if not category_counts.empty:

                    category_labels = [
                        f"{label} {int(value)} "
                        f"({value / category_counts.sum() * 100:.0f}%)"
                        for label, value
                        in zip(
                            category_counts.index,
                            category_counts.values
                        )
                    ]

                    fig_category = go.Figure(
                        data=[
                            go.Pie(
                                labels=category_labels,
                                values=category_counts.values,
                                hole=0.58,
                                textinfo="none",
                                domain=dict(
                                    x=[0.00, 0.55],
                                    y=[0.08, 0.92],
                                ),
                                hovertemplate=(
                                    "%{label}"
                                    "<extra></extra>"
                                ),
                            )
                        ]
                    )

                    fig_category.update_layout(
                        height=115,
                        margin=dict(
                            l=0,
                            r=0,
                            t=0,
                            b=0
                        ),
                        showlegend=True,
                        legend=dict(
                            orientation="v",
                            x=0.72,
                            y=0.5,
                            xanchor="left",
                            yanchor="middle",
                            font=dict(size=8),
                        ),
                        font=dict(size=8),
                    )

                    st.plotly_chart(
                        fig_category,
                        use_container_width=True,
                        config={
                            "displayModeBar": False
                        },
                        key="bf_moc_category_donut",
                    )

                else:

                    st.info(
                        "No MOC Category data found."
                    )

            else:

                st.info(
                    "MOC Category column not found."
                )

        # ----------------------------------------------------
        # MOC REGISTER
        # MOC No. / Description / Type / Status / Remarks
        # ----------------------------------------------------
        show_moc_register(moc)

# ============================================================
# ROW 2 — PSSR / INCIDENT / TRAINING
# ============================================================

a, b, c = st.columns(
    [1.20, 1.40, 1.40],
    gap="small"
)

# ============================================================
# 5 — PSSR
# ============================================================

with a:
    with st.container(border=True):
        show_module_title(
            5,
            "",
            "PRE-STARTUP SAFETY REVIEW (PSSR)"
        )

        x = status_counts(pssr)

        show_metric_row([
            ("TOTAL PSSR", x["total"]),
            ("COMPLETED", x["completed"]),
            ("PENDING", x["pending"]),
            ("OVERDUE", x["overdue"]),
        ])

        show_register(
            "PSSR REGISTER",
            pssr,
            [
                "PSSR No",
                "PSSR No.",
                "PSSR ID",
                "ID"
            ],
            [
                "PSSR Description",
                "Description",
                "PSSR Name"
            ],
            [
                "Status",
                "Current Status"
            ],
        )

# ============================================================
# 6 — PROCESS SAFETY INCIDENT
# ============================================================

with b:
    with st.container(
            border=True,
            height=460
    ):

        show_module_title(
            6,
            "⚠",
            "PROCESS SAFETY INCIDENT"
        )

        # ----------------------------------------------------
        # COLUMN MAPPING
        # B = Department → Total Incidents
        # F = Incident Classification
        # G = Incident Level
        # J = Investigation Status
        # ----------------------------------------------------

        if len(incident.columns) >= 2:
            department_col = incident.columns[1]
        else:
            department_col = None

        if len(incident.columns) >= 6:
            classification_col = incident.columns[5]
        else:
            classification_col = find_col(
                incident,
                [
                    "Incident Classification",
                    "Incident classification",
                    "Classification",
                ]
            )

        if len(incident.columns) >= 7:
            level_col = incident.columns[6]
        else:
            level_col = find_col(
                incident,
                [
                    "Incident Level",
                    "Incident level",
                    "Level",
                ]
            )

        if len(incident.columns) >= 10:
            investigation_status_col = (
                incident.columns[9]
            )
        else:
            investigation_status_col = find_col(
                incident,
                [
                    "Investigation status",
                    "Investigation Status",
                    "Investigation",
                    "Status",
                ]
            )

        # ----------------------------------------------------
        # TOTAL INCIDENTS — COLUMN B
        # ----------------------------------------------------

        total_incidents = 0

        if (
                department_col
                and not incident.empty
        ):
            department_values = (
                incident[department_col]
                .fillna("")
                .astype(str)
                .str.strip()
            )

            total_incidents = (
                department_values
                .replace(
                    [
                        "",
                        "-",
                        "nan",
                        "None"
                    ],
                    pd.NA
                )
                .notna()
                .sum()
            )

        # ----------------------------------------------------
        # INVESTIGATION STATUS — COLUMN J
        # ----------------------------------------------------

        investigation_completed = 0
        investigation_pending = 0

        if (
                investigation_status_col
                and not incident.empty
        ):
            investigation_values = (
                incident[
                    investigation_status_col
                ]
                .fillna("")
                .astype(str)
                .str.strip()
                .str.lower()
            )

            investigation_completed = (
                investigation_values
                .isin(
                    [
                        "completed",
                        "complete",
                        "closed",
                        "done",
                    ]
                )
                .sum()
            )

            investigation_pending = (
                investigation_values
                .isin(
                    [
                        "pending",
                        "ongoing",
                        "open",
                        "in progress",
                        "in-progress",
                    ]
                )
                .sum()
            )

        # ----------------------------------------------------
        # COMPACT KPI CARDS
        # ----------------------------------------------------

        m1, m2, m3 = st.columns(
            [0.70, 0.70, 0.70],
            gap="small"
        )

        with m1:

            st.metric(
                "TOTAL INCIDENTS",
                int(total_incidents)
            )

        with m2:

            st.metric(
                "INVESTIGATION COMPLETED",
                int(investigation_completed)
            )

        with m3:

            st.metric(
                "INVESTIGATION PENDING",
                int(investigation_pending)
            )

        # ====================================================
        # TWO PSI GRAPHS
        # ====================================================

        p1, p2 = st.columns(
            2,
            gap="small"
        )

        # ====================================================
        # CLASSIFICATION-WISE
        # ====================================================

        with p1:

            st.caption(
                "CLASSIFICATION-WISE DISTRIBUTION"
            )

            if (
                    classification_col
                    and not incident.empty
            ):

                classification_values = (
                    incident[
                        classification_col
                    ]
                    .fillna("")
                    .astype(str)
                    .str.strip()
                    .str.lower()
                )

                classification_values = (
                    classification_values[
                        ~classification_values.isin(
                            [
                                "",
                                "-",
                                "nan",
                                "none"
                            ]
                        )
                    ]
                )

                # ------------------------------------------------
                # STANDARDIZE CLASSIFICATION
                # Serious must be checked FIRST
                # ------------------------------------------------

                classification_values = (
                    classification_values.map(
                        lambda x:
                        "Serious Process Incident"
                        if (
                                "serious" in x
                                and "process" in x
                                and "incident" in x
                        )
                        else
                        "Process Incident"
                        if (
                                "process" in x
                                and "incident" in x
                        )
                        else
                        "Near Miss"
                        if (
                                "near" in x
                                and "miss" in x
                        )
                        else x.title()
                    )
                )

                classification_counts = (
                    classification_values
                    .value_counts()
                )

                classification_order = [
                    "Near Miss",
                    "Process Incident",
                    "Serious Process Incident",
                ]

                classification_counts = (
                    classification_counts
                    .reindex(
                        classification_order,
                        fill_value=0
                    )
                )

                classification_counts = (
                    classification_counts[
                        classification_counts > 0
                        ]
                )

                if not classification_counts.empty:

                    fig_class = go.Figure()

                    fig_class.add_trace(
                        go.Bar(
                            x=classification_counts.values,
                            y=classification_counts.index,
                            orientation="h",
                            width=0.50,
                            text=classification_counts.values,
                            textposition="outside",

                            marker=dict(
                                color=[
                                    "#0751a5"
                                    if x == "Near Miss"
                                    else "#e31b23"
                                    if x == "Process Incident"
                                    else "#808080"
                                    for x
                                    in classification_counts.index
                                ]
                            ),

                            hovertemplate=(
                                "%{y}: %{x}"
                                "<extra></extra>"
                            ),
                        )
                    )

                    fig_class.update_layout(
                        height=180,
                        margin=dict(
                            l=5,
                            r=25,
                            t=5,
                            b=5
                        ),
                        showlegend=False,
                        font=dict(size=9),

                        xaxis=dict(
                            title=None,
                            dtick=1,
                            showgrid=True,
                            gridcolor="#e1e7ef",
                            zeroline=False,
                        ),

                        yaxis=dict(
                            title=None,
                            autorange="reversed",
                        ),

                        plot_bgcolor="white",
                        paper_bgcolor="white",
                    )

                    st.plotly_chart(
                        fig_class,
                        use_container_width=True,
                        config={
                            "displayModeBar": False
                        },
                        key="psi_classification_chart",
                    )

                else:

                    st.info(
                        "No incident classification data."
                    )

            else:

                st.info(
                    "Incident Classification column not found."
                )

        # ====================================================
        # LEVEL-WISE
        # ====================================================

        with p2:

            st.caption(
                "LEVEL-WISE DISTRIBUTION"
            )

            if (
                    level_col
                    and not incident.empty
            ):

                level_values = (
                    incident[level_col]
                    .fillna("")
                    .astype(str)
                    .str.strip()
                    .str.lower()
                )

                level_values = (
                    level_values[
                        ~level_values.isin(
                            [
                                "",
                                "-",
                                "nan",
                                "none"
                            ]
                        )
                    ]
                )

                level_values = (
                    level_values.map(
                        lambda x:
                        "Level 1"
                        if "level 1" in x
                        else
                        "Level 2"
                        if "level 2" in x
                        else
                        "Level 3"
                        if "level 3" in x
                        else
                        "Level 4"
                        if "level 4" in x
                        else x.title()
                    )
                )

                level_counts = (
                    level_values
                    .value_counts()
                )

                level_order = [
                    "Level 1",
                    "Level 2",
                    "Level 3",
                    "Level 4",
                ]

                level_counts = (
                    level_counts
                    .reindex(
                        level_order,
                        fill_value=0
                    )
                )

                level_counts = (
                    level_counts[
                        level_counts > 0
                        ]
                )

                if not level_counts.empty:

                    fig_level = go.Figure()

                    fig_level.add_trace(
                        go.Bar(
                            x=level_counts.values,
                            y=level_counts.index,
                            orientation="h",
                            width=0.50,
                            text=level_counts.values,
                            textposition="outside",

                            marker=dict(
                                color=[
                                    "#0751a5"
                                    if x == "Level 1"
                                    else "#e31b23"
                                    if x == "Level 2"
                                    else "#808080"
                                    if x == "Level 3"
                                    else "#173f70"
                                    for x
                                    in level_counts.index
                                ]
                            ),

                            hovertemplate=(
                                "%{y}: %{x}"
                                "<extra></extra>"
                            ),
                        )
                    )

                    fig_level.update_layout(
                        height=180,
                        margin=dict(
                            l=5,
                            r=25,
                            t=5,
                            b=5
                        ),
                        showlegend=False,
                        font=dict(size=9),

                        xaxis=dict(
                            title=None,
                            dtick=1,
                            showgrid=True,
                            gridcolor="#e1e7ef",
                            zeroline=False,
                        ),

                        yaxis=dict(
                            title=None,
                            autorange="reversed",
                        ),

                        plot_bgcolor="white",
                        paper_bgcolor="white",
                    )

                    st.plotly_chart(
                        fig_level,
                        use_container_width=True,
                        config={
                            "displayModeBar": False
                        },
                        key="psi_level_chart",
                    )

                else:

                    st.info(
                        "No incident level data."
                    )

            else:

                st.info(
                    "Incident Level column not found."
                )

# ============================================================
# 7 — TRAINING
# ============================================================

with c:
    with st.container(
            border=True,
            height=460
    ):

        show_module_title(
            7,
            "♙",
            "TRAINING"
        )

        if (
                training is None
                or training.empty
        ):

            st.info(
                "No training data found."
            )

        else:

            tr = training.copy()

            # ------------------------------------------------
            # FIND COLUMNS
            # ------------------------------------------------

            process_col = find_col(
                tr,
                ["Process"]
            )

            department_col = find_col(
                tr,
                ["Departments"]
            )

            total_l08_col = find_col(
                tr,
                [
                    "Total Employees (L08 & Above)"
                ]
            )

            total_below_l08_col = find_col(
                tr,
                [
                    "Total Employees (Below L08)"
                ]
            )

            total_associates_col = find_col(
                tr,
                [
                    "Total Associates"
                ]
            )

            total_contractual_col = find_col(
                tr,
                [
                    "Total Contractual Workers"
                ]
            )

            completed_l08_col = find_col(
                tr,
                [
                    "Completed Training (L08 & Above)"
                ]
            )

            completed_below_l08_col = find_col(
                tr,
                [
                    "Completed Training (Below L08)"
                ]
            )

            completed_associates_col = find_col(
                tr,
                [
                    "Completed Training (Associates)"
                ]
            )

            completed_contractual_col = find_col(
                tr,
                [
                    "Completed Training (Contracts)"
                ]
            )

            required_columns = [
                process_col,
                total_l08_col,
                total_below_l08_col,
                total_associates_col,
                total_contractual_col,
                completed_l08_col,
                completed_below_l08_col,
                completed_associates_col,
                completed_contractual_col,
            ]

            if any(
                    col is None
                    for col in required_columns
            ):

                st.error(
                    "Training Google Sheet column mapping error."
                )

            else:

                def to_number(series):

                    return pd.to_numeric(
                        series
                        .astype(str)
                        .str.replace(
                            ",",
                            "",
                            regex=False
                        )
                        .str.replace(
                            "%",
                            "",
                            regex=False
                        )
                        .str.strip(),
                        errors="coerce"
                    )


                # ------------------------------------------------
                # BUILD HEATMAP DATA
                # ------------------------------------------------

                heatmap_df = pd.DataFrame()

                # ============================================================
                # TRAINING - ALL DEPARTMENTS CONSOLIDATED
                # ============================================================

                # Convert all employee/count columns to numbers
                tr["__TOTAL_L08"] = to_number(tr[total_l08_col])
                tr["__TOTAL_BELOW"] = to_number(tr[total_below_l08_col])
                tr["__TOTAL_ASSOC"] = to_number(tr[total_associates_col])
                tr["__TOTAL_CONTRACT"] = to_number(tr[total_contractual_col])

                tr["__DONE_L08"] = to_number(tr[completed_l08_col])
                tr["__DONE_BELOW"] = to_number(tr[completed_below_l08_col])
                tr["__DONE_ASSOC"] = to_number(tr[completed_associates_col])
                tr["__DONE_CONTRACT"] = to_number(tr[completed_contractual_col])

                # Combine ALL departments for each module
                summary = (
                    tr.groupby(process_col, dropna=True)[
                        [
                            "__TOTAL_L08",
                            "__TOTAL_BELOW",
                            "__TOTAL_ASSOC",
                            "__TOTAL_CONTRACT",
                            "__DONE_L08",
                            "__DONE_BELOW",
                            "__DONE_ASSOC",
                            "__DONE_CONTRACT",
                        ]
                    ]
                    .sum()
                    .reset_index()
                )

                # Module name
                summary["Module"] = (
                    summary[process_col]
                    .fillna("")
                    .astype(str)
                    .str.strip()
                )

                # Calculate consolidated percentage
                summary["L08 & Above"] = (
                    summary["__DONE_L08"]
                    .div(summary["__TOTAL_L08"].replace(0, float("nan")))
                    .mul(100)
                )

                summary["Below L08"] = (
                    summary["__DONE_BELOW"]
                    .div(summary["__TOTAL_BELOW"].replace(0, float("nan")))
                    .mul(100)
                )

                summary["Associates"] = (
                    summary["__DONE_ASSOC"]
                    .div(summary["__TOTAL_ASSOC"].replace(0, float("nan")))
                    .mul(100)
                )

                summary["Contractual"] = (
                    summary["__DONE_CONTRACT"]
                    .div(summary["__TOTAL_CONTRACT"].replace(0, float("nan")))
                    .mul(100)
                )

                # Remove blank modules
                summary = summary[
                    summary["Module"] != ""
                    ].copy()

                # Final table - SAME FORMAT
                heatmap_df = summary[
                    [
                        "Module",
                        "L08 & Above",
                        "Below L08",
                        "Associates",
                        "Contractual",
                    ]
                ].copy()

                heatmap_df = heatmap_df[
                    heatmap_df["Module"]
                    .str.strip() != ""
                    ].copy()

                # ------------------------------------------------
                # MODULE ORDER
                # ------------------------------------------------

                module_order = [
                    "PSM GA",
                    "PT",
                    "PHA",
                    "MOC",
                    "OP",
                    "BOWTIE",
                    "PSSR",
                    "LOPA",
                    "MIQA",
                ]

                heatmap_df["_order"] = (
                    heatmap_df["Module"]
                    .apply(
                        lambda x:
                        module_order.index(x)
                        if x in module_order
                        else 999
                    )
                )

                heatmap_df = (
                    heatmap_df
                    .sort_values(
                        ["_order", "Module"]
                    )
                    .drop(
                        columns="_order"
                    )
                )

                # ------------------------------------------------
                # DISPLAY DATA
                # ------------------------------------------------

                display_df = heatmap_df.copy()

                percentage_columns = [
                    "L08 & Above",
                    "Below L08",
                    "Associates",
                    "Contractual",
                ]

                for column in percentage_columns:
                    display_df[column] = (
                        display_df[column]
                        .apply(
                            lambda x:
                            "—"
                            if pd.isna(x)
                            else f"{x:.2f}%"
                        )
                    )


                # ------------------------------------------------
                # HEATMAP STYLE
                # ------------------------------------------------

                def heatmap_style(column):

                    styles = []

                    for value in column:

                        numeric = pd.to_numeric(
                            str(value)
                            .replace("%", ""),
                            errors="coerce"
                        )

                        if pd.isna(numeric):

                            styles.append(
                                "background-color:#ffffff;"
                                "color:#8a97a5;"
                                "text-align:center;"
                            )

                        else:

                            numeric = max(
                                0,
                                min(
                                    100,
                                    float(numeric)
                                )
                            )

                            if numeric <= 50:

                                ratio = numeric / 50

                                red = 255

                                green = int(
                                    220
                                    + (
                                            25 * ratio
                                    )
                                )

                                blue = int(
                                    220
                                    - (
                                            70 * ratio
                                    )
                                )

                            else:

                                ratio = (
                                                numeric - 50
                                        ) / 50

                                red = int(
                                    255
                                    - (
                                            55 * ratio
                                    )
                                )

                                green = 245

                                blue = int(
                                    150
                                    + (
                                            45 * ratio
                                    )
                                )

                            styles.append(
                                f"background-color:"
                                f"rgb({red},{green},{blue});"
                                "color:#173f70;"
                                "font-weight:700;"
                                "text-align:center;"
                            )

                    return styles


                styled_df = (
                    display_df.style
                    .apply(
                        heatmap_style,
                        subset=percentage_columns,
                        axis=0,
                    )
                    .set_properties(
                        subset=["Module"],
                        **{
                            "font-weight": "700",
                            "color": "#173f70",
                            "text-align": "left",
                        }
                    )
                    .set_properties(
                        **{
                            "font-size": "10px",
                            "border": "1px solid #d6e1eb",
                        }
                    )
                    .set_table_styles(
                        [
                            {
                                "selector": "th",
                                "props": [
                                    (
                                        "background-color",
                                        "#073f7c"
                                    ),
                                    (
                                        "color",
                                        "white"
                                    ),
                                    (
                                        "font-weight",
                                        "700"
                                    ),
                                    (
                                        "text-align",
                                        "center"
                                    ),
                                    (
                                        "font-size",
                                        "10px"
                                    ),
                                ],
                            }
                        ]
                    )
                )

                st.dataframe(
                    styled_df,
                    use_container_width=True,
                    hide_index=True,
                    height=245,
                    key="bf_training_heatmap",
                )

# ============================================================
# ROW 3 — SOC / SOL + AUDIT
# ============================================================

a, b = st.columns(
    [1.35, 1.65],
    gap="small"
)

# ============================================================
# 8 — SOC / SOL DEVIATION
# ============================================================

# ============================================================
# 8 — SOC / SOL DEVIATION
# ============================================================

with a:
    with st.container(border=True):

        show_module_title(
            8,
            "",
            "SOC / SOL DEVIATION"
        )

        if soc is None or soc.empty:

            st.info(
                "No SOC / SOL data available."
            )

        else:

            # ====================================================
            # COLUMN MAPPING
            # ====================================================

            month_col = find_col(
                soc,
                [
                    "Month",
                    "MONTH",
                    "month"
                ]
            )

            department_col = find_col(
                soc,
                [
                    "Department",
                    "DEPARTMENT",
                    "department"
                ]
            )

            soc_col = find_col(
                soc,
                [
                    "SOC Deviation Nos.",
                    "SOC Deviation No.",
                    "SOC Deviation",
                    "SOC"
                ]
            )

            sol_col = find_col(
                soc,
                [
                    "SOL Deviation Nos.",
                    "SOL Deviation No.",
                    "SOL Deviation",
                    "SOL"
                ]
            )

            if month_col is None:

                st.error(
                    "Month column not found in SOC / SOL Google Sheet."
                )

            else:

                df_socsol = soc.copy()

                # =================================================
                # CLEAN MONTH
                # =================================================

                df_socsol["_MONTH"] = (
                    df_socsol[month_col]
                    .fillna("")
                    .astype(str)
                    .str.strip()
                )

                df_socsol = df_socsol[
                    df_socsol["_MONTH"] != ""
                    ].copy()

                # =================================================
                # DEPARTMENT CLEANING
                # =================================================

                if department_col is not None:

                    df_socsol["_DEPARTMENT"] = (
                        df_socsol[department_col]
                        .fillna("")
                        .astype(str)
                        .str.strip()
                    )

                else:

                    df_socsol["_DEPARTMENT"] = "Unknown"

                # =================================================
                # SOC VALUE
                # =================================================

                if soc_col is not None:

                    df_socsol["_SOC_VALUE"] = (
                        pd.to_numeric(
                            df_socsol[soc_col],
                            errors="coerce"
                        )
                        .fillna(0)
                    )

                else:

                    df_socsol["_SOC_VALUE"] = 0

                # =================================================
                # SOL VALUE
                # =================================================

                if sol_col is not None:

                    df_socsol["_SOL_VALUE"] = (
                        pd.to_numeric(
                            df_socsol[sol_col],
                            errors="coerce"
                        )
                        .fillna(0)
                    )

                else:

                    df_socsol["_SOL_VALUE"] = 0

                # =================================================
                # DEPARTMENT LIST
                # =================================================

                departments = sorted(
                    [
                        x
                        for x in
                        df_socsol["_DEPARTMENT"]
                        .dropna()
                        .unique()
                        if str(x).strip() != ""
                    ]
                )

                department_options = [
                                         "All Departments"
                                     ] + departments

                # =================================================
                # FILTER DATA
                # =================================================

                if selected_department == "All Departments":

                    selected_socsol = df_socsol.copy()

                else:

                    selected_socsol = df_socsol[
                        df_socsol["_DEPARTMENT"]
                        .astype(str)
                        .str.strip()
                        == selected_department
                        ].copy()

                # =================================================
                # TOTAL SOC / SOL
                # =================================================

                total_soc = int(
                    selected_socsol["_SOC_VALUE"].sum()
                )

                total_sol = int(
                    selected_socsol["_SOL_VALUE"].sum()
                )

                # =================================================
                # KPI CARDS
                # =================================================

                k1, k2 = st.columns(
                    2,
                    gap="small"
                )

                with k1:

                    st.metric(
                        "TOTAL SOC",
                        total_soc
                    )

                with k2:

                    st.metric(
                        "TOTAL SOL",
                        total_sol
                    )

                # =================================================
                # FY MONTHS
                # =================================================

                fy_months = [
                    "April-26",
                    "May-26",
                    "June-26",
                    "July-26",
                    "August-26",
                    "September-26",
                    "October-26",
                    "November-26",
                    "December-26",
                    "January-27",
                    "February-27",
                    "March-27",
                ]

                # =================================================
                # MONTH-WISE SOC
                # =================================================

                soc_monthly = (
                    selected_socsol
                    .groupby(
                        "_MONTH",
                        as_index=False
                    )["_SOC_VALUE"]
                    .sum()
                )

                # =================================================
                # MONTH-WISE SOL
                # =================================================

                sol_monthly = (
                    selected_socsol
                    .groupby(
                        "_MONTH",
                        as_index=False
                    )["_SOL_VALUE"]
                    .sum()
                )

                # =================================================
                # CREATE FY TABLE
                # =================================================

                monthly = pd.DataFrame(
                    {
                        "_MONTH": fy_months
                    }
                )

                monthly = monthly.merge(
                    soc_monthly,
                    on="_MONTH",
                    how="left"
                )

                monthly = monthly.merge(
                    sol_monthly,
                    on="_MONTH",
                    how="left"
                )

                monthly["_SOC_VALUE"] = (
                    monthly["_SOC_VALUE"]
                    .fillna(0)
                )

                monthly["_SOL_VALUE"] = (
                    monthly["_SOL_VALUE"]
                    .fillna(0)
                )

                # =================================================
                # GRAPH
                # =================================================

                fig = go.Figure()

                # -------------------------------------------------
                # SOC
                # -------------------------------------------------

                fig.add_trace(
                    go.Scatter(
                        x=monthly["_MONTH"],
                        y=monthly["_SOC_VALUE"],
                        mode="lines+markers+text",
                        name="SOC Deviation",

                        text=(
                            monthly["_SOC_VALUE"]
                            .astype(int)
                        ),

                        textposition="top center",

                        line=dict(
                            width=3
                        ),

                        marker=dict(
                            size=7
                        ),

                        hovertemplate=(
                            "<b>SOC</b><br>"
                            "Month: %{x}<br>"
                            "Deviation: %{y}"
                            "<extra></extra>"
                        )
                    )
                )

                # -------------------------------------------------
                # SOL
                # -------------------------------------------------

                fig.add_trace(
                    go.Scatter(
                        x=monthly["_MONTH"],
                        y=monthly["_SOL_VALUE"],
                        mode="lines+markers+text",
                        name="SOL Deviation",

                        text=(
                            monthly["_SOL_VALUE"]
                            .astype(int)
                        ),

                        textposition="bottom center",

                        line=dict(
                            width=3,
                            dash="solid"
                        ),

                        marker=dict(
                            size=7
                        ),

                        hovertemplate=(
                            "<b>SOL</b><br>"
                            "Month: %{x}<br>"
                            "Deviation: %{y}"
                            "<extra></extra>"
                        )
                    )
                )

                # =================================================
                # GRAPH FORMAT
                # =================================================

                fig.update_layout(

                    height=280,

                    margin=dict(
                        l=45,
                        r=20,
                        t=45,
                        b=45
                    ),

                    title=dict(
                        text=(
                            "MONTH-WISE DISTRIBUTION "
                            "OF SOC / SOL DEVIATION"
                        ),
                        x=0.5,
                        xanchor="center",
                        font=dict(
                            size=12,
                            color="#173f73"
                        )
                    ),

                    xaxis=dict(
                        title="Month",
                        categoryorder="array",
                        categoryarray=fy_months,
                        tickangle=-45,
                        showgrid=False
                    ),

                    yaxis=dict(
                        title="No. of Deviations",
                        rangemode="tozero",
                        showgrid=True,
                        dtick=1
                    ),

                    legend=dict(
                        orientation="v",
                        x=0.78,
                        y=1.28,
                        xanchor="left",
                        yanchor="top",
                        font=dict(size=9),
                        bgcolor="rgba(255,255,255,0)"
                    ),

                    plot_bgcolor="white",
                    paper_bgcolor="white",

                    hovermode="x unified"
                )

                # =================================================
                # DISPLAY GRAPH
                # =================================================

                st.plotly_chart(
                    fig,
                    use_container_width=True,
                    config={
                        "displayModeBar": False
                    },
                    key="all_department_soc_sol_chart"
                )

# ============================================================
# 9 — AUDIT / COMPLIANCE
# ============================================================

with b:
    with st.container(
            border=True,
            height=430
    ):

        show_module_title(
            9,
            "♙",
            "AUDIT / COMPLIANCE"
        )

        # ========================================================
        # AUDIT KPI SUMMARY
        # ========================================================

        # TOTAL NO. OF AUDIT DONE = records with a valid Audit Date
        # PENDING FOR AUDIT = records without a valid Audit Date
        total_audit_done = 0
        pending_audit = 0

        if audit is not None and not audit.empty:
            audit_date_for_kpi = find_col(
                audit,
                [
                    "Audit Date",
                    "Last Audit Date",
                    "Date",
                ],
            )

            if audit_date_for_kpi:
                audit_dates = pd.to_datetime(
                    audit[audit_date_for_kpi],
                    errors="coerce",
                )
                total_audit_done = int(audit_dates.notna().sum())
                pending_audit = int(audit_dates.isna().sum())
            else:
                pending_audit = int(len(audit))

        ak1, ak2 = st.columns(2, gap="small")

        with ak1:
            st.metric("TOTAL NO. OF AUDIT DONE", total_audit_done)

        with ak2:
            st.metric("PENDING FOR AUDIT", pending_audit)

        # ========================================================
        # AUDIT / COMPLIANCE REGISTER
        # ========================================================

        st.markdown(
            '<div class="section-title">AUDIT / COMPLIANCE REGISTER</div>',
            unsafe_allow_html=True,
        )

        if audit is None or audit.empty:
            st.info("No audit records available.")
        else:
            audit_sno_col = find_col(
                audit,
                [
                    "S.No.",
                    "S.No",
                    "Sr No",
                    "Sr. No",
                    "Serial No",
                    "S No",
                ],
            )

            audit_dept_col = find_col(
                audit,
                [
                    "Department",
                    "Dept",
                    "Department Name",
                ],
            )

            audit_date_col = find_col(
                audit,
                [
                    "Audit Date",
                    "Last Audit Date",
                    "Date",
                ],
            )

            audit_score_col = find_col(
                audit,
                [
                    "Audit Score",
                    "Score",
                    "Audit Compliance",
                    "Compliance",
                    "Compliance %",
                    "Percentage",
                ],
            )

            register_df = pd.DataFrame(index=audit.index)

            if audit_sno_col:
                register_df["S.No."] = audit[audit_sno_col].fillna("").astype(str).str.strip()
            else:
                register_df["S.No."] = range(1, len(audit) + 1)

            if audit_dept_col:
                register_df["Department"] = audit[audit_dept_col].fillna("").astype(str).str.strip()
            else:
                register_df["Department"] = ""

            if audit_date_col:
                register_df["Audit Date"] = audit[audit_date_col].fillna("").astype(str).str.strip()
            else:
                register_df["Audit Date"] = ""

            if audit_score_col:
                register_df["Audit Score"] = audit[audit_score_col].fillna("").astype(str).str.strip()
            else:
                register_df["Audit Score"] = ""

            # IMPORTANT:
            # Use the real URL extracted from Google Sheet HTML.
            # Do not use the displayed report name as a URL.
            if "__COMPLIANCE_REPORT_URL" in audit.columns:
                register_df["Compliance Report"] = (
                    audit["__COMPLIANCE_REPORT_URL"]
                    .fillna("")
                    .astype(str)
                    .str.strip()
                )
            else:
                register_df["Compliance Report"] = ""

            # --------------------------------------------------------
            # COMPLIANCE REPORT BUTTONS
            # --------------------------------------------------------
            # The button uses the live URL extracted from the Google Sheet.
            # No report URL is fixed in this code.
            rows_html = []

            for _, row in register_df.iterrows():
                report_url = str(row.get("Compliance Report", "")).strip()
                if report_url.startswith("http://") or report_url.startswith("https://"):
                    report_cell = (
                        f'<a href="{report_url}" target="_blank" rel="noopener noreferrer" '
                        'style="display:inline-block;padding:4px 10px;'
                        'background:#07518b;color:#ffffff !important;border-radius:4px;'
                        'text-decoration:none;font-weight:700;font-size:9px;">'
                        'VIEW REPORT ↗</a>'
                    )
                else:
                    report_cell = '<span style="color:#9aa7b3;font-size:9px;">Not attached</span>'

                rows_html.append(
                    f"""
                    <tr>
                        <td>{row['S.No.']}</td>
                        <td>{row['Department']}</td>
                        <td>{row['Audit Date']}</td>
                        <td>{row['Audit Score']}</td>
                        <td>{report_cell}</td>
                    </tr>
                    """
                )

            # Use a real HTML component so the live SharePoint/Google-sheet
            # hyperlink remains an actual clickable button.
            audit_table_html = f"""
            <!DOCTYPE html>
            <html>
            <head>
            <style>
                html, body {{
                    margin:0; padding:0; background:transparent;
                    font-family:Arial, Helvetica, sans-serif;
                }}
                .audit-register-wrap {{
                    width:100%; max-height:400px; overflow-y:auto;
                    border:1px solid #d5e0e8; border-radius:4px;
                    background:#ffffff;
                }}
                table {{
                    width:100%; border-collapse:collapse; table-layout:fixed;
                    font-size:9px; color:#173f70;
                }}
                th {{
                    position:sticky; top:0; z-index:2;
                    background:#f3f7fa; color:#627689;
                    font-weight:800; text-align:left;
                    padding:5px 8px; border-bottom:1px solid #d5e0e8;
                    line-height:1.1;
                }}
                td {{
                    padding:5px 8px; border-bottom:1px solid #e1e8ee;
                    line-height:1.1;
                    white-space:nowrap; overflow:hidden; text-overflow:ellipsis;
                }}
                th:nth-child(1), td:nth-child(1) {{ width:8%; }}
                th:nth-child(2), td:nth-child(2) {{ width:24%; }}
                th:nth-child(3), td:nth-child(3) {{ width:18%; }}
                th:nth-child(4), td:nth-child(4) {{ width:17%; }}
                th:nth-child(5), td:nth-child(5) {{ width:33%; }}
                .report-btn {{
                    display:inline-block; padding:4px 10px;
                    background:#07518b; color:#ffffff !important;
                    border-radius:4px; text-decoration:none !important;
                    font-weight:700; font-size:9px; cursor:pointer;
                }}
                .report-btn:hover {{ background:#063e70; }}
            </style>
            </head>
            <body>
                <div class="audit-register-wrap">
                    <table>
                        <thead>
                            <tr>
                                <th>S.No.</th>
                                <th>Department</th>
                                <th>Audit Date</th>
                                <th>Audit Score</th>
                                <th>Compliance Report</th>
                            </tr>
                        </thead>
                        <tbody>
                            {''.join(rows_html)}
                        </tbody>
                    </table>
                </div>
            </body>
            </html>
            """

            st.components.v1.html(
                audit_table_html,
                height=305,
                scrolling=False,
            )